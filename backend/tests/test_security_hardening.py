from datetime import datetime,timezone
from hashlib import sha256
from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

from fastapi import HTTPException,Request,Response
from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest
from sqlalchemy import delete,select

from app.api import routes
from app.core.config import Settings
from app.core.database import SessionLocal
from app.main import app,frontend_candidate
from app.models.entities import AuthSession,BusinessCalendar,Classification,Email,EmailStatus,LoginThrottle,Mailbox,Role,SLATier,SecurityAuditEvent,User
from app.services import classification,jobs
from app.services.graph import validate_graph_url
from app.services.oauth_transactions import consume_transaction,oauth_cookie_name,save_transaction
from app.services.secrets import get_mailbox_token,set_mailbox_token


ADMIN_EMAIL="admin@oeis.local"
ADMIN_PASSWORD="OEIS-Admin@July2026#47"


def admin_login(client:TestClient):
    response=client.post("/api/auth/login",data={"username":ADMIN_EMAIL,"password":ADMIN_PASSWORD});assert response.status_code==200
    return response,{"Authorization":f"Bearer {response.json()['access_token']}"}


def test_login_uses_httponly_cookie_and_refresh_rotation_detects_replay():
    with TestClient(app) as client:
        login,_=admin_login(client);body=login.json();cookie_name=routes._refresh_cookie_name();old=client.cookies.get(cookie_name)
        assert set(body)=={"access_token","token_type","expires_in"} and old
        cookie=login.headers["set-cookie"].lower();assert "httponly" in cookie and "samesite=strict" in cookie
        refreshed=client.post("/api/auth/refresh");assert refreshed.status_code==200
        current=client.cookies.get(cookie_name);assert current and current!=old
        client.cookies.set(cookie_name,old,path="/");assert client.post("/api/auth/refresh").status_code==401
        client.cookies.set(cookie_name,current,path="/");assert client.post("/api/auth/refresh").status_code==401


def test_refresh_rejects_cross_site_origin():
    with TestClient(app) as client:
        admin_login(client)
        assert client.post("/api/auth/refresh",headers={"Origin":"https://attacker.example","Sec-Fetch-Site":"cross-site"}).status_code==403


def test_login_throttle_is_shared_in_database_and_errors_are_generic():
    suffix=uuid4().hex;email=f"throttle-{suffix}@example.com";password="Throttle-Test-Password-123!"
    with SessionLocal() as db:db.add(User(email=email,name="Throttle",password_hash=routes.hash_password(password),role=Role.MANAGER,active=True));db.commit()
    try:
        with TestClient(app) as client:
            first=client.post("/api/auth/login",data={"username":email,"password":"wrong-password-value"});assert first.status_code==401
            blocked=client.post("/api/auth/login",data={"username":email,"password":password});assert blocked.status_code==429 and "Retry-After" in blocked.headers
            unknown=client.post("/api/auth/login",data={"username":f"missing-{suffix}@example.com","password":"wrong-password-value"});assert unknown.status_code==429 or unknown.json()["detail"]==first.json()["detail"]
    finally:
        with SessionLocal() as db:db.execute(delete(LoginThrottle));db.commit()


def test_mailbox_refresh_tokens_are_authenticated_encrypted_and_tamper_detected():
    suffix=uuid4().hex
    with SessionLocal() as db:
        mailbox=Mailbox(address=f"cipher-{suffix}@example.com",display_name="Cipher",provider="gmail",timezone="UTC");db.add(mailbox);db.flush();set_mailbox_token(mailbox,"provider-refresh-secret");db.commit();db.refresh(mailbox)
        assert mailbox.graph_refresh_token is None and "provider-refresh-secret" not in mailbox.token_ciphertext
        assert get_mailbox_token(mailbox)=="provider-refresh-secret"
        mailbox.token_ciphertext=("A" if mailbox.token_ciphertext[0]!="A" else "B")+mailbox.token_ciphertext[1:];db.flush()
        with pytest.raises(Exception):get_mailbox_token(mailbox)


def test_oauth_transaction_is_browser_session_provider_bound_and_one_time():
    suffix=uuid4().hex;state=f"state-{suffix}";binding=f"binding-{suffix}"
    with TestClient(app) as client:
        admin_login(client)
        with SessionLocal() as db:
            user=db.scalar(select(User).where(User.email==ADMIN_EMAIL));session=db.scalar(select(AuthSession).where(AuthSession.user_id==user.id).order_by(AuthSession.created_at.desc()))
            mailbox=Mailbox(address=f"oauth-{suffix}@example.com",display_name="OAuth",provider="gmail",timezone="UTC");db.add(mailbox);db.flush()
            save_transaction(db,Response(),state,binding,"gmail",user.id,session.id,mailbox.id,{"code_verifier":"pkce-verifier"});db.commit()
            def request(cookie_value:str)->Request:
                cookie=f"{oauth_cookie_name()}={cookie_value}".encode()
                return Request({"type":"http","method":"GET","path":"/callback","query_string":b"","headers":[(b"cookie",cookie)],"scheme":"https","server":("testserver",443),"client":("127.0.0.1",1)})
            with pytest.raises(HTTPException):consume_transaction(db,request("wrong-browser"),state,"gmail")
            with pytest.raises(HTTPException):consume_transaction(db,request(binding),state,"microsoft")
            _,payload=consume_transaction(db,request(binding),state,"gmail");assert payload["code_verifier"]=="pkce-verifier";db.commit()
            with pytest.raises(HTTPException):consume_transaction(db,request(binding),state,"gmail")


def test_manager_mailbox_scope_is_deny_by_default_and_filters_all_views():
    suffix=uuid4().hex;manager_email=f"scope-{suffix}@example.com";password="Manager-Scope-Password-123!"
    with TestClient(app) as client:
        _,admin_headers=admin_login(client)
        manager=client.post("/api/users",headers=admin_headers,json={"email":manager_email,"name":"Scoped Manager","password":password,"role":"manager","active":True}).json()
        allowed=client.post("/api/mailboxes",headers=admin_headers,json={"address":f"allowed-{suffix}@example.com","display_name":"Allowed","timezone":"UTC"}).json()
        denied=client.post("/api/mailboxes",headers=admin_headers,json={"address":f"denied-{suffix}@example.com","display_name":"Denied","timezone":"UTC"}).json()
        client.put(f"/api/users/{manager['id']}/mailbox-access",headers=admin_headers,json={"mailbox_ids":[allowed["id"]]})
        with SessionLocal() as db:
            now=datetime.now(timezone.utc);rows=[]
            for mailbox_id,label in ((allowed["id"],"allowed"),(denied["id"],"denied")):
                rows.append(Email(mailbox_id=mailbox_id,message_id=f"scope-{label}-{suffix}",sender=f"{label}-{suffix}@customer.example",receiver="support@example.com",subject=f"Scope {label} {suffix}",received_time=now,folder="inbox",categories=[],classification=Classification.CUSTOMER,status=EmailStatus.PENDING,pending_hours=30,sla_tier=SLATier.CRITICAL))
            db.add_all(rows);db.add_all([BusinessCalendar(mailbox_id=None,timezone="UTC"),BusinessCalendar(mailbox_id=allowed["id"],timezone="UTC"),BusinessCalendar(mailbox_id=denied["id"],timezone="UTC")]);db.commit();denied_email=rows[1].id
        manager_login=client.post("/api/auth/login",data={"username":manager_email,"password":password});manager_headers={"Authorization":f"Bearer {manager_login.json()['access_token']}"}
        options=client.get("/api/mailbox-options",headers=manager_headers).json();assert [row["id"] for row in options]==[allowed["id"]]
        pending=client.get(f"/api/emails/pending?search={suffix}",headers=manager_headers).json();assert pending["total"]==1 and pending["items"][0]["subject"]==f"Scope allowed {suffix}"
        assert client.get(f"/api/emails/{denied_email}",headers=manager_headers).status_code==404
        calendars=client.get("/api/settings/business-calendars",headers=manager_headers).json();assert {row["mailbox_id"] for row in calendars}=={None,allowed["id"]}
        report=client.get("/api/reports/daily?dimension=customer",headers=manager_headers).json();assert any(suffix in row["dimension"] and row["dimension"].startswith("allowed") for row in report["rows"]);assert not any(row["dimension"].startswith(f"denied-{suffix}") for row in report["rows"])


def test_xlsx_export_neutralizes_formula_cells():
    suffix=uuid4().hex;malicious=f"=HYPERLINK(\"https://attacker.invalid/{suffix}\")"
    with TestClient(app) as client:
        _,headers=admin_login(client);mailbox=client.post("/api/mailboxes",headers=headers,json={"address":f"xlsx-{suffix}@example.com","display_name":"XLSX","timezone":"UTC"}).json()
        with SessionLocal() as db:db.add(Email(mailbox_id=mailbox["id"],message_id=f"xlsx-{suffix}",sender=malicious,receiver="support@example.com",subject="Formula test",received_time=datetime.now(timezone.utc),folder="inbox",categories=[],classification=Classification.CUSTOMER,status=EmailStatus.PENDING,pending_hours=1,sla_tier=SLATier.HEALTHY));db.commit()
        response=client.get("/api/reports/daily/export?dimension=customer&format=xlsx",headers=headers);assert response.status_code==200
        sheet=load_workbook(BytesIO(response.content),data_only=False).active
        cells=[cell for row in sheet.iter_rows() for cell in row if suffix in str(cell.value)]
        assert cells and all(cell.data_type!="f" and str(cell.value).startswith("'") for cell in cells)


def test_html_notifications_escape_email_metadata():
    with TestClient(app):
        with SessionLocal() as db:
            html=jobs._summary_html(db,"<img src=x onerror=alert(1)>",[SimpleNamespace(sender="<b>sender</b>",subject="<script>alert(1)</script>")],1,1,"1 Hr","https://oeis.example/?x=<bad>")
    assert "<script>" not in html and "<img" not in html and "&lt;script&gt;" in html and "&lt;img" in html


def test_regex_timeout_fails_safe(monkeypatch):
    def timeout(*args,**kwargs):raise TimeoutError
    monkeypatch.setattr(classification.regex,"search",timeout)
    rule=classification.Rule(1,"subject","(a+)+$",Classification.IGNORE)
    assert classification.classify("customer@example.com","a"*2048+"!",{},[rule])==Classification.CUSTOMER


@pytest.mark.parametrize("url",["http://graph.microsoft.com/v1.0/me","https://evil.example/v1.0/me","https://graph.microsoft.com.evil.example/v1.0/me","https://user@graph.microsoft.com/v1.0/me","https://graph.microsoft.com:444/v1.0/me"])
def test_graph_url_allowlist_rejects_ssrf_destinations(url):
    with pytest.raises(RuntimeError):validate_graph_url(url,Settings())
    assert validate_graph_url("https://graph.microsoft.com/v1.0/me",Settings()).startswith("https://graph.microsoft.com/")


def test_security_headers_no_store_and_admin_audit_events_exist():
    with TestClient(app) as client:
        login,headers=admin_login(client);dashboard=client.get("/api/dashboard/kpis",headers=headers)
        assert dashboard.headers["cache-control"]=="no-store" and dashboard.headers["x-frame-options"]=="DENY" and "default-src 'self'" in dashboard.headers["content-security-policy"]
        events=client.get("/api/security/audit-events",headers=headers);assert events.status_code==200 and any(row["action"]=="auth.login" for row in events.json())
        assert "refresh_token" not in login.text


def test_frontend_fallback_cannot_traverse_outside_distribution_directory(tmp_path,monkeypatch):
    outside=tmp_path/"outside-secret.txt";outside.write_text("must-not-be-served")
    monkeypatch.setattr("app.main.dist",tmp_path/"public")
    (tmp_path/"public").mkdir();(tmp_path/"public"/"index.html").write_text("index")
    assert frontend_candidate("../outside-secret.txt")==tmp_path/"public"/"index.html"


def test_production_configuration_rejects_insecure_defaults():
    settings=Settings(environment="production",database_url="mssql://example",frontend_url="https://oeis.example",dashboard_url="https://oeis.example",graph_redirect_uri="https://oeis.example/api/graph/oauth/callback",google_redirect_uri="https://oeis.example/api/gmail/oauth/callback",enable_api_docs=False)
    with pytest.raises(RuntimeError,match="JWT_SECRET"):settings.validate_security()
